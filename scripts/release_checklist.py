#!/usr/bin/env python3
"""Generate the MANUAL pre-release review checklist as an .xlsx workbook.

    python scripts/release_checklist.py            # -> validation/release-checklist.xlsx
    python scripts/release_checklist.py --out X    # write somewhere else

This is the eyes-and-hands gate. The automated gates (`--selftest`, the F3 corpus,
`--livetest`, `--motiontest`) prove the renderer computes the right numbers; they cannot
tell you the window opened, the menus are reachable, the palette looks like a palette, or
the app survives being resized. Everything in here needs a human.

⚠KEEP IT HONEST. Every row names a control that EXISTS: the fractal list, colour methods,
palettes, menu items and panel sections below were read out of the source, not invented. If
a row cannot be performed because the feature moved, that is a FAIL of this document — fix
the row, do not skip it.

The workbook has two sheets:
  * `Run`      — the checklist itself: Step, Area, Action, Expected, Actual, Pass/Fail, Notes.
  * `Cover`    — build under test, tester, platform, GPU, and the sign-off summary.

Pass/Fail is a dropdown (PASS / FAIL / BLOCKED / N-A) so a completed run can be filtered.
"""

import argparse
import datetime
import os
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Read out of the source so the script and the app cannot silently drift apart.
FRACTALS = [
    "Mandelbrot", "Multibrot3", "Multibrot4", "Multibrot5", "Tricorn",
    "Burning Ship", "Celtic", "Buffalo", "Phoenix", "Newton",
]
COLOR_METHODS = ["Smooth", "Stripe", "Triangle inequality", "Orbit trap", "Distance", "Decomposition"]
PALETTES = ["Ember", "Ice", "Nebula", "Grayscale"]

# (Area, Action, Expected)
STEPS = [
    # ---------------------------------------------------------------- launch
    ("Launch",
     "Launch the release build with a CLEAN profile: rename %APPDATA%\\Fractadyne\\Fractadyne\\config "
     "to config.bak first, so this is a true first run.",
     "Window opens within a few seconds. No console errors. No crash dialog."),
    ("Launch",
     "Read the title bar.",
     "Shows 'Fractadyne v<version> (build <n>)' and the version matches the release being tested."),
    ("Launch",
     "On first run, the welcome / quick-start dialog appears. Read it, then dismiss it.",
     "Dialog is readable, no clipped text, buttons work, and it closes without reappearing."),
    ("Launch",
     "Observe the initial view.",
     "The full Mandelbrot set is drawn, centred, correctly proportioned, and finishes rendering "
     "(no permanent blank, black, or half-drawn frame)."),

    # ---------------------------------------------------------------- layout
    ("Layout & theme",
     "Inspect the overall window: menu bar, toolbar, fractal area, right-hand Controls panel, status bar.",
     "All five regions present. Nothing overlaps, is cut off, or spills outside the window."),
    ("Layout & theme",
     "Inspect text and icon rendering across the menu bar, toolbar and panel.",
     "Text is crisp and fully legible at the current DPI. Icons are not blurry, clipped or missing "
     "(no placeholder boxes)."),
    ("Layout & theme",
     "Check colour scheme consistency: panel background, headings, buttons, highlighted/active items.",
     "Consistent theme throughout. Sufficient contrast to read every label. Active/selected states "
     "are visually distinct."),
    ("Layout & theme",
     "Read the status bar at the bottom.",
     "Shows centre coordinates, cursor position, zoom, and iteration count. Values are populated, "
     "not blank or placeholder."),
    ("Layout & theme",
     "Expand every Controls panel section: Navigate, Coloring, Quality, Effects, Overlays, Advanced, "
     "About <fractal>, Performance.",
     "All sections expand and collapse. No empty sections. Controls are aligned and fully visible."),
    ("Layout & theme",
     "Collapse the Controls panel (View > Control panel, or 'Hide control panel'), then re-enable it.",
     "Panel hides, the fractal area reflows to use the space, and re-enabling restores it with "
     "settings intact."),

    # ---------------------------------------------------------------- sizing
    ("Window sizing",
     "Drag the window edge to make the window much NARROWER, then much WIDER.",
     "The fractal re-renders to the new aspect ratio without stretching or distortion. The panel "
     "and status bar stay usable. No crash."),
    ("Window sizing",
     "Drag to make the window very SHORT (a wide letterbox), then very TALL.",
     "Layout still holds; no controls become unreachable; the image stays correctly proportioned."),
    ("Window sizing",
     "Maximise the window, then restore it.",
     "Renders at the new size both ways. No lingering black bands or stale image."),
    ("Window sizing",
     "Resize the window rapidly back and forth for ~5 seconds.",
     "App stays responsive and recovers to a correct, complete frame. No crash, no permanently "
     "blank view."),
    ("Window sizing",
     "If you have a second monitor at a different DPI, drag the window to it.",
     "UI rescales and stays legible; the image re-renders correctly."),

    # ---------------------------------------------------------------- navigation
    ("Navigation",
     "Click once in the fractal area with 'Click to zoom' enabled.",
     "View zooms in centred on the clicked point by the selected Factor. Motion is smooth and "
     "lands where you clicked."),
    ("Navigation",
     "Try each zoom Factor: 2x, 4x, 10x, 50x, 100x.",
     "Each factor visibly changes the zoom step. The status-bar zoom value increases accordingly."),
    ("Navigation",
     "Use the toolbar magnifier buttons to zoom in and out several times.",
     "Zoom in and out both work and are centred on the view."),
    ("Navigation",
     "Drag with the mouse to pan around the set.",
     "The image follows the cursor 1:1 while dragging and resolves to a sharp frame when released."),
    ("Navigation",
     "Scroll the mouse wheel forward and back over the image.",
     "Wheel zooms in/out smoothly about the cursor."),
    ("Navigation",
     "Press Ctrl+Z (or Backspace) several times, then Ctrl+Y (or Shift+Backspace).",
     "Undo steps back through the previous views; redo returns forward. Coordinates match what "
     "you visited."),
    ("Navigation",
     "Click the Home toolbar button ('Zoom out to full view').",
     "Returns to the whole set, animated, ending at the standard home view."),
    ("Navigation",
     "Use View > Reset to default view.",
     "Instantly returns to the default view for the current fractal."),
    ("Navigation",
     "Enable View > Minimap overview and navigate around.",
     "Minimap appears, shows the current viewport indicator, and tracks your navigation."),

    # ---------------------------------------------------------------- deep zoom
    ("Deep zoom",
     "Pick a filament edge and zoom continuously to about 1e6x. Watch the image the whole way in.",
     "Detail keeps resolving. No banding artefacts, no blocky tiles left behind, no all-black frames."),
    ("Deep zoom",
     "Continue to roughly 1e15x, then 1e30x. Watch the Performance panel 'mode' field.",
     "Mode changes as depth increases (direct -> perturb df32 -> floatexp). Each transition is "
     "seamless; the image does not flash to garbage or flat colour."),
    ("Deep zoom",
     "Continue to about 1e100x, pausing to let each view fully settle.",
     "Image still resolves to clean structure. Status bar precision and orbit length grow. No "
     "speckle field, no solid-colour frame."),
    ("Deep zoom",
     "Push on to 1e300x or deeper if patience allows.",
     "Still renders real structure. Deep zoom is the headline feature; a blank or noise frame here "
     "is a release blocker."),
    ("Deep zoom",
     "At depth, confirm 'Normalize deep colors' (Coloring) is ON and look at the exterior banding.",
     "The exterior shows smooth colour bands, NOT per-pixel salt-and-pepper speckle."),
    ("Deep zoom",
     "Toggle 'Normalize deep colors' OFF at that same deep view.",
     "The exterior becomes visibly noisier/aliased - which confirms the setting is actually doing "
     "something. Turn it back ON."),
    ("Deep zoom",
     "While deep, drag to pan and then let the view settle.",
     "Panning stays responsive, and the frame resolves fully afterwards rather than staying coarse."),
    ("Deep zoom",
     "Zoom back out to home from maximum depth in one continuous motion.",
     "No freeze, no crash, no device-loss dialog. The app remains responsive throughout."),

    # ---------------------------------------------------------------- fractals
] + [
    ("Fractal types",
     "Fractal menu > %s. Let it render." % name,
     "Correct, recognisable %s renders at a sensible default view. No blank frame, no error." % name)
    for name in FRACTALS
] + [
    ("Fractal types",
     "With a formula that supports it selected, zoom in ~1e6x on an interesting edge.",
     "Detail resolves correctly for that formula, not just for Mandelbrot."),
    ("Fractal types",
     "Return to Mandelbrot and enable Fractal > Julia mode.",
     "A Julia set renders for the current c. Image is plausible and complete."),
    ("Fractal types",
     "Enable Fractal > Dual view (Mandelbrot <-> Julia), then move the cursor over the left pane.",
     "Window splits. Left shows the parameter set; right shows the Julia set for the cursor's c "
     "and updates as you move."),
    ("Fractal types",
     "Zoom into the Julia pane while in dual view.",
     "The Julia side zooms independently and resolves correctly."),
    ("Fractal types",
     "Turn off dual view and Julia mode.",
     "Returns cleanly to the single Mandelbrot view."),

    # ---------------------------------------------------------------- colour
] + [
    ("Coloring",
     "Coloring > Method: %s. Observe the image." % m,
     "Colouring visibly changes and produces a coherent image (no all-black, all-white, or "
     "uniform flat frame)." )
    for m in COLOR_METHODS
] + [
    ("Coloring",
     "Coloring > Palette: cycle through %s." % ", ".join(PALETTES),
     "Each palette applies immediately and looks distinct from the others."),
    ("Coloring",
     "Drag the Cycle slider across its range.",
     "Band frequency changes smoothly and live. No flicker, no stuck frame."),
    ("Coloring",
     "Drag the Offset slider across its range.",
     "Colours shift phase smoothly through the palette."),
    ("Coloring",
     "With 'Normalize deep colors' ON, toggle 'Log color scale'.",
     "Colour distribution changes noticeably; image stays coherent. (Log scale acts on the "
     "NORMALIZED mapping - with normalization off it has no visible effect.)"),
    ("Coloring",
     "Open 'Edit gradient…', change a colour stop, and apply.",
     "Editor opens, the edit is reflected in the live image, and the dialog closes cleanly."),
    ("Coloring",
     "Set Animate to on and watch for a few seconds, then set it back to Off.",
     "Palette animates smoothly at the set speed and stops cleanly when disabled."),
    ("Coloring",
     "In Effects, toggle 'Binary (set)' and 'Duotone'.",
     "Each visibly changes the rendering and can be turned back off."),
    ("Coloring",
     "In Effects, enable '3D relief lighting' and drag the Light angle control.",
     "Relief shading appears and the highlight direction follows the angle."),
    ("Coloring",
     "In Effects, enable 'Distance glow' and adjust its strength/width.",
     "Glow appears around structure edges and responds to the controls."),

    # ---------------------------------------------------------------- quality
    ("Quality",
     "In Quality, toggle 'Auto-scale iterations with zoom' off and set Iterations manually to a low "
     "value (e.g. 200) at a moderately deep view.",
     "Image visibly loses detail / gains flat interior - i.e. the control takes effect."),
    ("Quality",
     "Raise Iterations to a high value (e.g. 100000) at that same view.",
     "Detail returns. Render takes longer but completes; app stays responsive."),
    ("Quality",
     "Re-enable 'Auto-scale iterations with zoom'.",
     "Iteration count in the status bar starts tracking zoom depth again."),
    ("Quality",
     "Change Anti-alias between 1x and 2x (and higher if offered) and compare edges.",
     "Higher AA visibly smooths edges. Frame time increases accordingly in the Performance panel."),
    ("Quality",
     "Open the Performance panel and watch it while navigating.",
     "FPS, frame time, GPU time, mode, effective iterations, precision and orbit length all update "
     "and look plausible."),

    # ---------------------------------------------------------------- locations
    ("Locations",
     "Navigate > Go to location… and enter a known deep coordinate.",
     "Dialog accepts full-precision input and the view jumps exactly there (status bar matches)."),
    ("Locations",
     "Navigate > Random location, several times.",
     "Each jump lands somewhere valid and renders real structure - not a blank or all-interior frame."),
    ("Locations",
     "Navigate > Bookmarks > Add current view, then navigate away, then restore it from the menu.",
     "Bookmark is saved with a usable name and restores the exact view."),
    ("Locations",
     "Navigate > Bookmarks… - delete a bookmark, including one you have just added.",
     "Dialog opens, deletion works, the remaining list and its thumbnails stay consistent. "
     "(There is no rename: a bookmark is named when it is added.)"),
    ("Locations",
     "Navigate > Share location… and copy the result.",
     "Produces a shareable string/link for the current view without error."),
    ("Locations",
     "Navigate > Import .kfr… with a Kalles Fraktaler / Fraktaler-3 parameter file.",
     "Import succeeds and the view lands at the imported coordinates."),
    ("Locations",
     "File > Open view or location… with a previously exported PNG (or .fdn).",
     "The stored view is recovered and rendered."),
    ("Locations",
     "File > Gallery… - browse and open an entry.",
     "Gallery opens, thumbnails display, and selecting one loads that view."),

    # ---------------------------------------------------------------- output
    ("Export & capture",
     "Press Ctrl+S (Snapshot) at an interesting view.",
     "An image is written without freezing the UI, and the app reports where it went."),
    ("Export & capture",
     "Open the saved snapshot in an image viewer.",
     "It matches what was on screen (same framing and colours) and is not corrupt or truncated."),
    ("Export & capture",
     "File > Export image… at a LARGER size than the window (e.g. 4K) with supersampling on.",
     "Progress is shown, the export completes, and the app stays responsive or clearly indicates "
     "it is working."),
    ("Export & capture",
     "Open the 4K export.",
     "Full resolution, complete image, no missing tiles, no seams between tiles, no colour banding "
     "differences across tile boundaries."),
    ("Export & capture",
     "Export once at a DEEP view (1e30x or deeper).",
     "Completes and matches the on-screen framing and colouring."),

    # ---------------------------------------------------------------- tools
    ("Tools",
     "Press M (Tools > Find minibrot + zoom to it).",
     "The app locates a minibrot and animates to it, landing on a real structure."),
    ("Tools",
     "Tools > Newton / Misiurewicz solver… - run it on the current view.",
     "Solver dialog opens, runs, and reports a result without hanging."),
    ("Tools",
     "Tools > Play tour… and play a bundled tour for at least a minute.",
     "Camera moves smoothly through the script, frames resolve, captions (if any) display correctly."),
    ("Tools",
     "Close the tour player.",
     "Playback stops and normal interactive control returns."),
    ("Tools",
     "Tools > Tour from current view… ",
     "Creates a tour seeded at the current view without error."),
    ("Tools",
     "Tools > Benchmark… and let it finish.",
     "Benchmark runs to completion and reports results; app remains usable afterwards."),

    # ---------------------------------------------------------------- help/settings
    ("Help & settings",
     "Press F1 (Help & reference).",
     "Help window opens, content is readable and scrolls, no broken layout or missing sections."),
    ("Help & settings",
     "Help > Diagnostics… ",
     "Diagnostics window opens and shows real values (GPU/adapter, paths, log location), "
     "including a 'Deep-zoom arithmetic' line naming the backend in use."),
    ("Help & settings",
     "Help > About (the last Help section). Read the 'Deep-zoom arithmetic' line.",
     "Names the arithmetic that has actually run and what the build contains. On the standard "
     "build: astro-float. On the accelerated build: rug, with the MPFR/GMP versions. Before any "
     "deep render it may say none has run yet - that is correct, not a missing value."),
    ("Help & settings",
     "Help > Faster deep zoom… ",
     "On the STANDARD build: explains the optional accelerated download and offers two buttons. "
     "'Download for this version' opens a URL containing THIS version's tag; 'All releases' opens "
     "the releases page. On the ACCELERATED build: says you are already running it instead, and "
     "offers no download."),
    ("Help & settings",
     "Help > Report an issue… ",
     "Opens the issue reporting path correctly (browser or dialog) with the app's details."),
    ("Help & settings",
     "Help > Check for updates.",
     "Performs a check and reports a clear result (up to date, or an update is available). No hang."),
    ("Help & settings",
     "File > Settings - change a setting, close and reopen the menu.",
     "Setting persists and takes effect."),

    # ------------------------------------------------------- accelerated build (optional artifact)
    # Mark this whole block N-A if you are only reviewing the standard download.
    ("Accelerated build",
     "Extract fractadyne-<tag>-windows-x64-accelerated.zip to a NEW folder and run fractadyne.exe "
     "from it. Do this on a machine (or account) with no MSYS2 and no MinGW on PATH.",
     "Window opens normally. NO 'the code execution cannot proceed' or missing-DLL dialog - the "
     "package must be self-contained apart from the .dll files beside the executable."),
    ("Accelerated build",
     "Help > About on that build.",
     "'Deep-zoom arithmetic' names rug and reports the MPFR and GMP versions."),
    ("Accelerated build",
     "Check that your existing locations, bookmarks and last session are present.",
     "All there. Settings live in the user profile, not beside the executable, so the two builds "
     "share them and nothing needs importing."),
    ("Accelerated build",
     "Open the same deep location (1e50 or deeper) in BOTH builds and compare the images "
     "side by side; export a PNG from each if unsure.",
     "Visually identical. The two backends are byte-identical by construction and by test, so ANY "
     "visible difference is a bug worth stopping the release for."),
    ("Accelerated build",
     "Dive into a deep view on both builds and watch the pause before the picture starts "
     "resolving (the reference-orbit build).",
     "Noticeably shorter on the accelerated build. This is the only user-visible difference there "
     "should be. Note both timings."),
    ("Accelerated build",
     "Delete libmpfr-6.dll from the accelerated folder and try to launch it. Restore it after.",
     "Fails to start with a missing-DLL error rather than silently falling back to the slow "
     "arithmetic. Restoring the file makes it work again."),

    # ---------------------------------------------------------------- persistence
    ("Persistence",
     "Navigate to a distinctive deep view, then quit via File > Quit.",
     "App exits cleanly - no hang, no crash dialog, no lingering process."),
    ("Persistence",
     "Relaunch the app.",
     "It reopens on the SAME view you left, with the same fractal, palette and settings."),
    ("Persistence",
     "Check config\\logs for crash-*.txt files created during this run.",
     "No new crash reports. (If any exist, this run FAILS and the file must be attached.)"),

    # ---------------------------------------------------------------- stability
    ("Stability",
     "Leave the app running idle at a deep view for 5 minutes.",
     "No creeping memory growth to the point of instability, no watchdog restart, view still correct."),
    ("Stability",
     "Do a fast, sustained zoom-in / zoom-out / pan session for ~2 minutes.",
     "No freeze, no device-loss restart, no crash. The app keeps up or degrades gracefully."),
    ("Stability",
     "Switch fractal type, colour method and palette rapidly several times in a row.",
     "No crash and no stuck frame; the final selection is what renders."),
    ("Sign-off",
     "Review every FAIL and BLOCKED row above with the release decision in mind.",
     "Either all rows PASS, or each non-PASS has an agreed decision (fix before release / accept "
     "and document). Record the decision on the Cover sheet."),
]



# ---------------------------------------------------------------------------------------------
# Which machine check enforces each step, in STEP ORDER. See design/checklist-automation.md.
#
# Paired POSITIONALLY rather than carried inside the tuples because STEPS is partly generated
# (one row per fractal, one per colour method), so not every step is a literal to edit. The area
# is repeated here so a reordered or inserted step misaligns loudly instead of silently crediting
# one step with another's coverage - `scripts/checklist_coverage.py` checks both.
#
#   planned:<rest>   declared but NOT yet implemented; counts as outstanding, never as covered
#   test:<name>      a cargo test name substring, must resolve against `cargo test -- --list`
#   selftest:<name>  a --selftest check name, must appear in selftest.rs
#   uitest:<name>    a --uitest check name, must appear in uitest.rs
#   harness:<flag>   a CLI flag that must exist in the argument parser
#   script:<path>    a repo file that must exist
#   partial:<e>      the effect is enforced; a gesture or an aesthetic judgement is not
#   manual/process   no machine enforcer is possible - the reason is in the plan document
ENFORCERS = [
    ("Launch", "planned:uitest:clean-launch-no-crash"),
    ("Launch", "test:title_string_matches_version"),
    ("Launch", "test:welcome_shows_once"),
    ("Launch", "selftest:home"),
    ("Layout & theme", "planned:uitest:layout-regions"),
    ("Layout & theme", "partial:test:every_ui_glyph_has_a_font_that_can_draw_it"),
    ("Layout & theme", "planned:partial:test:theme_contrast_meets_minimum"),
    ("Layout & theme", "planned:uitest:status-bar-populated"),
    ("Layout & theme", "planned:partial:uitest:panel-sections"),
    ("Layout & theme", "planned:test:panel_toggle_reflows"),
    ("Window sizing", "planned:uitest:aspect-ratio"),
    ("Window sizing", "planned:uitest:aspect-ratio"),
    ("Window sizing", "planned:uitest:maximize-restore"),
    ("Window sizing", "planned:uitest:rapid-resize"),
    ("Window sizing", "manual"),
    ("Navigation", "partial:test:click_zoom_applies_factor"),
    ("Navigation", "partial:test:click_zoom_applies_factor"),
    ("Navigation", "partial:test:toolbar_zoom_actions"),
    ("Navigation", "partial:test:pan_pixels_moves_exactly_one_pixel_per_pixel"),
    ("Navigation", "partial:test:zoom_at_keeps_the_point_under_the_cursor_fixed"),
    ("Navigation", "test:undo_redo_round_trip"),
    ("Navigation", "test:home_view_is_the_default"),
    ("Navigation", "test:reset_view_is_the_default"),
    ("Navigation", "partial:test:minimap_drag_signs"),
    ("Deep zoom", "partial:selftest:depth-ladder-coherent"),
    ("Deep zoom", "selftest:zoom-sequence across direct→df32 seam"),
    ("Deep zoom", "selftest:depth-ladder-coherent"),
    ("Deep zoom", "selftest:extreme-depth-coherent"),
    ("Deep zoom", "partial:selftest:normalize-reduces-speckle"),
    ("Deep zoom", "partial:selftest:normalize-reduces-speckle"),
    ("Deep zoom", "partial:selftest:pan consistency"),
    ("Deep zoom", "harness:--autodive-home"),
    ("Fractal types", "selftest:home"),
    ("Fractal types", "selftest:multibrot3"),
    ("Fractal types", "selftest:multibrot4"),
    ("Fractal types", "selftest:multibrot5"),
    ("Fractal types", "selftest:tricorn"),
    ("Fractal types", "selftest:burning-ship"),
    ("Fractal types", "selftest:celtic"),
    ("Fractal types", "selftest:buffalo"),
    ("Fractal types", "selftest:phoenix"),
    ("Fractal types", "selftest:newton"),
    ("Fractal types", "selftest:multibrot3-1e6"),
    ("Fractal types", "selftest:julia-coherent"),
    ("Fractal types", "partial:test:dual_view_splits_the_viewports"),
    ("Fractal types", "partial:test:julia_viewport_zooms_independently"),
    ("Fractal types", "test:leaving_dual_restores_the_single_view"),
    ("Coloring", "selftest:colour methods are all different"),
    ("Coloring", "selftest:colour methods are all different"),
    ("Coloring", "selftest:colour methods are all different"),
    ("Coloring", "selftest:colour methods are all different"),
    ("Coloring", "selftest:colour methods are all different"),
    ("Coloring", "selftest:colour methods are all different"),
    ("Coloring", "selftest:palettes are all different and coherent"),
    ("Coloring", "selftest:control changes the image"),
    ("Coloring", "selftest:control changes the image"),
    ("Coloring", "selftest:log-scale-changes-the-image"),
    ("Coloring", "partial:selftest:gradient-edit-changes-the-image"),
    ("Coloring", "partial:test:palette_animation_advances_and_stops"),
    ("Coloring", "selftest:control changes the image"),
    ("Coloring", "partial:selftest:control changes the image"),
    ("Coloring", "partial:selftest:control changes the image"),
    ("Quality", "partial:selftest:explicit iteration count honoured verbatim"),
    ("Quality", "partial:selftest:explicit iteration count honoured verbatim"),
    ("Quality", "test:recommended_max_iter_never_decreases_with_depth"),
    ("Quality", "selftest:supersampling softens edges"),
    ("Quality", "planned:partial:uitest:perf-fields-populated"),
    ("Locations", "test:go_to_round_trips_a_deep_coordinate"),
    ("Locations", "selftest:random-locations-coherent"),
    ("Locations", "test:bookmark_round_trip"),
    ("Locations", "test:bookmark_delete_keeps_the_list_consistent"),
    ("Locations", "test:share_string_round_trip"),
    ("Locations", "test:kfr_import_keeps_every_digit_of_a_deep_centre"),
    ("Locations", "selftest:metadata round-trips a deep view"),
    ("Locations", "partial:test:gallery_scan_and_load"),
    ("Export & capture", "partial:test:snapshot_writes_a_file"),
    ("Export & capture", "selftest:snapshot-matches-the-view"),
    ("Export & capture", "selftest:export-4k-complete"),
    ("Export & capture", "selftest:tiled chunked export is bit-identical"),
    ("Export & capture", "partial:selftest:deep-export-matches-the-view"),
    ("Tools", "partial:selftest:deep minibrot: size, framing, center accuracy"),
    ("Tools", "selftest:Misiurewicz multiplier vs closed forms"),
    ("Tools", "harness:--livetest"),
    ("Tools", "test:stopping_playback_restores_interaction"),
    ("Tools", "selftest:generated dive script round-trips"),
    ("Tools", "harness:--benchmark-std"),
    ("Help & settings", "planned:partial:uitest:help-sections-nonempty"),
    ("Help & settings", "planned:uitest:diagnostics-populated"),
    ("Help & settings", "test:about_names_the_running_backend"),
    ("Help & settings", "test:accelerated_asset_url"),
    ("Help & settings", "test:issue_url_is_well_formed"),
    ("Help & settings", "test:update_check_reaches_a_verdict"),
    ("Help & settings", "test:a_setting_survives_save_and_reload"),
    ("Accelerated build", "script:scripts/build-accelerated.ps1"),
    ("Accelerated build", "test:about_names_the_running_backend"),
    ("Accelerated build", "test:config_lives_in_the_user_profile"),
    ("Accelerated build", "partial:test:identity_holds_where_the_multiply_algorithms_diverge"),
    ("Accelerated build", "harness:--bench-bignum"),
    ("Accelerated build", "manual"),
    ("Persistence", "planned:test:clean_exit_leaves_no_process"),
    ("Persistence", "test:a_deep_view_and_its_colouring_survive_a_restart"),
    ("Persistence", "planned:uitest:no-crash-files"),
    ("Stability", "planned:harness:soak-liveness"),
    ("Stability", "harness:--torture"),
    ("Stability", "partial:selftest:rapid-switching-settles-on-the-final-choice"),
    ("Sign-off", "process"),
]

assert len(ENFORCERS) == len(STEPS), (
    "ENFORCERS has %d rows but STEPS has %d - every step needs one, even if it is 'manual'"
    % (len(ENFORCERS), len(STEPS))
)
for _i, ((_area, _enf), _step) in enumerate(zip(ENFORCERS, STEPS), start=1):
    assert _area == _step[0], (
        "enforcer row %d says area %r but step %d is in %r - the tables have drifted apart"
        % (_i, _area, _i, _step[0])
    )

def build(out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    # ---------------------------------------------------------------- Cover
    cov = wb.active
    cov.title = "Cover"
    title = Font(bold=True, size=16)
    lab = Font(bold=True)
    cov["A1"] = "Fractadyne — manual pre-release review"
    cov["A1"].font = title
    cov["A2"] = ("Walk the Run sheet top to bottom on the exact build you intend to publish. "
                 "Fill Actual result for anything that is not a clean pass.")
    cov["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    cov.merge_cells("A2:D2")
    cov.row_dimensions[2].height = 30

    rows = [
        ("Version under test", ""),
        ("Artifact under test (standard / accelerated)", ""),
        ("Build number", ""),
        ("Commit / tag", ""),
        ("Tester", ""),
        ("Date", ""),
        ("OS / version", ""),
        ("GPU / driver", ""),
        ("Display scaling (DPI)", ""),
        ("", ""),
        ("Steps passed", ""),
        ("Steps failed", ""),
        ("Steps blocked / N-A", ""),
        ("", ""),
        ("Release decision (SHIP / HOLD)", ""),
        ("Decision rationale", ""),
    ]
    r = 4
    for k, v in rows:
        if k:
            cov.cell(row=r, column=1, value=k).font = lab
            cov.cell(row=r, column=2, value=v)
        r += 1
    cov.column_dimensions["A"].width = 32
    cov.column_dimensions["B"].width = 52

    cov.cell(row=r + 1, column=1,
             value="Automated gates that must ALSO be green for this release "
                   "(these are not manual steps):").font = lab
    for i, g in enumerate([
        "fractadyne --selftest            -> exit 0. At 0.2.40: 140/140 checks, 18/18 goldens.",
        "   (The run prints its own totals. A DIFFERENT total is not automatically a failure,",
        "    but it must be explained - a silently skipped check looks exactly like a pass.)",
        "python validation/corpus/generate_corpus.py --check  -> 38/38 MATCH",
        "fractadyne --livetest tours/grand-tour.toml --size 480x270  -> 24/24, 0 drifted",
        "fractadyne --motiontest          -> VERDICT PASS",
        "cargo build --release            -> warning-free",
    ]):
        cov.cell(row=r + 2 + i, column=1, value=g)

    # ---------------------------------------------------------------- Run
    ws = wb.create_sheet("Run")
    headers = ["Step", "Area", "Action to take", "Expected result", "Automated by",
               "Actual result", "Pass/Fail", "Notes"]
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    area_fill = PatternFill("solid", fgColor="D9E2F3")
    top_wrap = Alignment(wrap_text=True, vertical="top")

    row = 2
    step = 1
    last_area = None
    for (area, action, expected), (_, enforcer) in zip(STEPS, ENFORCERS):
        if area != last_area:
            ws.cell(row=row, column=1, value=area).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = area_fill
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = area_fill
                ws.cell(row=row, column=c).border = border
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            row += 1
            last_area = area
        ws.cell(row=row, column=1, value=step).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row, column=2, value=area).alignment = top_wrap
        ws.cell(row=row, column=3, value=action).alignment = top_wrap
        ws.cell(row=row, column=4, value=expected).alignment = top_wrap
        ws.cell(row=row, column=5, value=enforcer).alignment = top_wrap
        ws.cell(row=row, column=6, value="").alignment = top_wrap
        ws.cell(row=row, column=7, value="").alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row, column=8, value="").alignment = top_wrap
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = border
        row += 1
        step += 1

    last_row = row - 1
    widths = {"A": 6, "B": 16, "C": 58, "D": 58, "E": 34, "F": 30, "G": 11, "H": 24}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), last_row)

    # DERIVE the verdict column from the header order. It was hard-coded as F, and adding the
    # "Automated by" column moved Pass/Fail to G - which would have left the dropdown on the
    # wrong column and the Cover sheet counting an empty one, with the workbook still opening
    # perfectly and the sign-off totals silently reading zero.
    verdict_col = get_column_letter(headers.index("Pass/Fail") + 1)

    dv = DataValidation(type="list", formula1='"PASS,FAIL,BLOCKED,N-A"', allow_blank=True, showDropDown=False)
    dv.error = "Choose PASS, FAIL, BLOCKED or N-A."
    ws.add_data_validation(dv)
    dv.add("{c}2:{c}{r}".format(c=verdict_col, r=last_row))

    # Cover formulas that count the run once it is filled in.
    cov["B13"] = '=COUNTIF(Run!{c}:{c},"PASS")'.format(c=verdict_col)
    cov["B14"] = '=COUNTIF(Run!{c}:{c},"FAIL")'.format(c=verdict_col)
    cov["B15"] = ('=COUNTIF(Run!{c}:{c},"BLOCKED")+COUNTIF(Run!{c}:{c},"N-A")'
                  .format(c=verdict_col))

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return step - 1, last_row


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.join(ROOT, "validation", "release-checklist.xlsx"))
    a = ap.parse_args()
    try:
        n, rows = build(a.out)
    except ImportError:
        sys.exit("openpyxl is required: python -m pip install openpyxl")
    print("wrote %s — %d steps, %d rows" % (a.out, n, rows))


if __name__ == "__main__":
    main()
